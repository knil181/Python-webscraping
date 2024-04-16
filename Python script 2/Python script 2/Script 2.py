import os
from openpyxl import load_workbook
from openpyxl import Workbook

# Step 1: Create a new xlsx document (master file)
master_wb = Workbook()
master_ws = master_wb.active

# ask user for the folders path which contains the excel file
folder = input("Input the path the path to the file: ")

# Step 2-6: Process files in the folder
folder_path = folder
start_row = 1

# Helper function to find the last row in a specific column of a worksheet
def find_last_row_in_column(worksheet, column):
    for row in range(worksheet.max_row, 1, -1):
        if worksheet.cell(row=row, column=column).value:
            return row
    return 1

# Helper function to get the last row in the master file's column H
def get_last_row_in_master(column):
    for row in range(master_ws.max_row, 1, -1):
        if master_ws.cell(row=row, column=column).value:
            return row
    return 1

for filename in os.listdir(folder_path):
    if filename.endswith(".xlsx"):
        file_path = os.path.join(folder_path, filename)
        wb = load_workbook(file_path)

        sheet = wb.worksheets[1]
        num_rows = sheet.max_row - 1

        if len(wb.worksheets) >= 2:
            # Step 3: Open the first sheet and copy data from cell A2 to G2 to the master file's cell number A2 to G1
            data_row_sheet1 = sheet = wb.worksheets[0][2]  # Row containing data A2 to G2 in the first sheet
            for _ in range(num_rows):
                for cell in data_row_sheet1:
                    master_ws.cell(row=start_row, column=cell.column).value = cell.value
                start_row += 1

            sheet2 = wb.worksheets[1]  # Second sheet
            last_row_in_column_H = get_last_row_in_master(8)  # Assuming column H is used for pasting data
            for row in sheet2.iter_rows(min_row=2, min_col=5, max_col=7, values_only=True):
                master_ws.cell(row=last_row_in_column_H + 1, column=8).value = row[0]
                master_ws.cell(row=last_row_in_column_H + 1, column=9).value = row[1]
                master_ws.cell(row=last_row_in_column_H + 1, column=10).value = row[2]
                last_row_in_column_H += 1

# Shifting data in Columns H to J one cell up
for column in range(8, 11):  # Columns H to J
    for row in range(1, master_ws.max_row):
        master_ws.cell(row=row, column=column).value = master_ws.cell(row=row+1, column=column).value

# Delete data from the last cell of Columns H to J
for column in range(8, 11):  # Columns H to J
    master_ws.cell(row=master_ws.max_row, column=column).value = None



#Shift all the data of the master file one cell down, So that the 1st row get's empty for me to write the header.
for row in range(master_ws.max_row, 1, -1):
    for column in range(1, master_ws.max_column + 1):
        master_ws.cell(row=row, column=column).value = master_ws.cell(row=row - 1, column=column).value

# Clear the first row (header row)
for column in range(1, master_ws.max_column + 1):
    master_ws.cell(row=1, column=column).value = None

# Add the header to the first row
header_data = ["Team A", "Team B", "bookmaker", "link to game", "odds for 1", "odds for X", "odds for 2", "Line",
               "odds for home team line", "odds for away team line"]
for col_idx, header in enumerate(header_data, start=1):
    master_ws.cell(row=1, column=col_idx).value = header

master_wb.save("master_file.xlsx")
